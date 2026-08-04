"""Dynamic Excel Template Generator for Marks Import.

Generates custom Excel workbooks based on the active AssessmentTemplate's
question components.
"""
from __future__ import annotations

import io
from typing import List

from app.features.marks_import.models import AssessmentTemplate

_HEADER_FILL = "2B3A66"


def build_dynamic_marks_template(template: AssessmentTemplate) -> bytes:
    """Build dynamic Excel workbook for a given AssessmentTemplate.

    If components exist (e.g. A, B, C, D), columns are:
        Student Roll Number | A | B | C | D

    If no components exist (e.g. Seminar, Open Book, Lab):
        Student Roll Number | Marks
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Marks Import"

    components = template.components_json or []

    # Build header row
    headers = ["Student Roll Number"]
    if components:
        for comp in components:
            headers.append(comp.get("key") or comp.get("label") or "Q")
    else:
        headers.append("Marks")

    sheet.append(headers)

    # Add sample rows
    if components:
        # Sample row 1
        sample1 = ["23BQ1A5401"]
        sample2 = ["23BQ1A5402"]
        for comp in components:
            max_m = float(comp.get("max_marks", 5.0))
            sample1.append(str(round(max_m * 0.9, 1)))
            sample2.append(str(round(max_m * 0.75, 1)))
        sheet.append(sample1)
        sheet.append(sample2)
    else:
        total_m = float(template.total_max_marks or 30.0)
        sheet.append(["23BQ1A5401", str(round(total_m * 0.85, 1))])
        sheet.append(["23BQ1A5402", str(round(total_m * 0.70, 1))])

    # Style sheet
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)

    for idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_letter = get_column_letter(idx)
        sheet.column_dimensions[col_letter].width = 22 if idx == 1 else 14

    sheet.row_dimensions[1].height = 22
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
