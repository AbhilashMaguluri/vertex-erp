"""Dynamic Excel parser for Marks Import."""
from __future__ import annotations

import io
import logging
from typing import Dict, List, Tuple

from app.core.exceptions import ValidationError
from app.features.marks_import.models import AssessmentTemplate
from app.features.marks_import.schemas import ParsedMarksRow
from app.features.marks_import.validation_service import validate_structure

logger = logging.getLogger("app.marks_import.parser")

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv")
MAX_UPLOAD_BYTES = 12 * 1024 * 1024


def parse_marks_excel(
    filename: str,
    content: bytes,
    template: AssessmentTemplate,
) -> Tuple[List[ParsedMarksRow], List[str]]:
    """Parse Excel file into dynamic ParsedMarksRow objects based on AssessmentTemplate."""
    if not content:
        raise ValidationError("The uploaded file is empty.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValidationError(f"File is {len(content) / 1_048_576:.1f} MB, above limit.")

    grid = _load_grid(filename, content)
    if not grid:
        raise ValidationError("This file is empty.")

    headers, header_index = _find_header(grid, template)
    if headers is None:
        raise ValidationError("Could not find a valid header row in the uploaded file.")

    column_map, structural_errors = validate_structure(headers, template)
    if structural_errors:
        return [], structural_errors

    rows: List[ParsedMarksRow] = []
    roll_col = column_map["student_roll"]
    components = template.components_json or []

    for offset, raw_row in enumerate(grid[header_index + 1:], start=header_index + 2):
        cells = [_cell_to_text(c) for c in raw_row]

        if not any(c.strip() for c in cells):
            continue

        student_roll = cells[roll_col].strip() if roll_col < len(cells) else ""

        question_scores: Dict[str, float] = {}
        total_marks: float | None = None

        if components:
            for comp in components:
                key = comp["key"]
                col_idx = column_map.get(key)
                val_str = cells[col_idx].strip() if (col_idx is not None and col_idx < len(cells)) else ""
                val_float = _parse_float(val_str)
                if val_float is not None:
                    question_scores[key] = val_float
        else:
            col_idx = column_map.get("marks")
            val_str = cells[col_idx].strip() if (col_idx is not None and col_idx < len(cells)) else ""
            total_marks = _parse_float(val_str)

        rows.append(ParsedMarksRow(
            row_number=offset,
            student_roll=student_roll,
            question_scores=question_scores,
            total_marks=total_marks,
        ))

    if not rows:
        raise ValidationError("No data rows found underneath the header row.")

    logger.info("Parsed %d marks rows from '%s'", len(rows), filename)
    return rows, []


def _parse_float(val: str) -> float | None:
    if not val:
        return None
    try:
        return float(val)
    except ValueError:
        return None


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _load_grid(filename: str, content: bytes) -> List[List]:
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return _load_csv(content)
    if lowered.endswith(".xls"):
        return _load_xls(content)
    if lowered.endswith((".xlsx", ".xlsm")):
        return _load_xlsx(content)
    raise ValidationError(f"Unsupported file type: {filename}")


def _load_xlsx(content: bytes) -> List[List]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("openpyxl not installed.") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(f"Could not open as Excel workbook: {exc}") from exc

    sheet = wb.active
    grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    wb.close()
    return grid


def _load_xls(content: bytes) -> List[List]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValidationError("xlrd not installed.") from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise ValidationError(f"Could not open as Excel workbook: {exc}") from exc

    sheet = book.sheet_by_index(0)
    grid: List[List] = []
    for r in range(sheet.nrows):
        grid.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
    return grid


def _load_csv(content: bytes) -> List[List]:
    import csv

    text = None
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:
        raise ValidationError("Could not determine CSV encoding.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


def _find_header(grid: List[List], template: AssessmentTemplate) -> Tuple[List[str] | None, int]:
    roll_aliases = {
        "student roll number", "roll number", "roll no", "rollno",
        "student roll", "registration number", "regd no", "roll", "ht no",
    }
    for idx, row in enumerate(grid[:15]):
        headers = [_cell_to_text(c).strip().lower() for c in row]
        if any(h in roll_aliases for h in headers):
            return [_cell_to_text(c).strip() for c in row], idx
    return None, -1
