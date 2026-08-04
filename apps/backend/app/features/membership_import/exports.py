"""Downloadable artefacts for the Membership Import feature.

Generates:
- Import error report (Import_Errors.xlsx)
- Credentials workbook
- Sample template
"""
from __future__ import annotations

import io
from typing import List, Sequence

from app.features.membership_import.models import MembershipImportBatch
from app.features.membership_import.schemas import (
    GeneratedStudentCredential,
    MembershipImportResultRecord,
    ValidationErrorRow,
)


_HEADER_FILL = "2B3A66"
_ERROR_FILL = "E1634B"


def _style_sheet(worksheet, column_count: int, widths: Sequence[int]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for idx in range(1, column_count + 1):
        cell = worksheet.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.column_dimensions[get_column_letter(idx)].width = (
            widths[idx - 1] if idx - 1 < len(widths) else 18
        )
    worksheet.row_dimensions[1].height = 22
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def build_error_report(validation_errors: List[ValidationErrorRow]) -> bytes:
    """Build Import_Errors.xlsx from validation errors."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Import Errors"

    sheet.append(["Row", "Error", "Description", "Suggested Fix"])

    for err in validation_errors:
        sheet.append([err.row, err.error, err.description, err.suggested_fix])

    _style_sheet(sheet, 4, [8, 40, 50, 40])

    # Highlight error cells
    error_font = Font(color="CC0000")
    for row_idx in range(2, len(validation_errors) + 2):
        for col_idx in range(1, 5):
            sheet.cell(row=row_idx, column=col_idx).font = error_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_credentials_workbook(
    batch: MembershipImportBatch,
    credentials: List[GeneratedStudentCredential],
) -> bytes:
    """Build the credentials download workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Student Credentials"

    sheet.append([
        "Roll Number", "Student Name", "Username",
        "Temporary Password", "Login Email", "Counselor", "Status",
    ])
    for cred in credentials:
        sheet.append([
            cred.roll_number, cred.full_name, cred.username,
            cred.temporary_password, cred.email, cred.counselor_email, cred.status,
        ])
    _style_sheet(sheet, 7, [18, 26, 18, 20, 30, 26, 12])

    # Highlight password column
    for idx in range(2, len(credentials) + 2):
        sheet.cell(row=idx, column=4).font = Font(bold=True, color="CC0000")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_report_workbook(
    batch: MembershipImportBatch,
    records: List[MembershipImportResultRecord],
    warnings: List[str],
    errors: List[str],
) -> bytes:
    """Build a detailed import report workbook."""
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Import Report"

    sheet.append(["Type", "Identifier", "Name", "Status", "Message", "Row"])
    for r in records:
        sheet.append([
            r.record_type, r.identifier,
            r.display_name or "", r.status,
            r.message or "", r.source_row_number or "",
        ])
    _style_sheet(sheet, 6, [14, 30, 26, 14, 50, 8])

    if warnings or errors:
        notes = wb.create_sheet("Warnings & Errors")
        notes.append(["Type", "Message"])
        for w in warnings:
            notes.append(["WARNING", w])
        for e in errors:
            notes.append(["ERROR", e])
        _style_sheet(notes, 2, [12, 80])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_sample_template() -> bytes:
    """Build the sample three-column template."""
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Membership Import"

    sheet.append(["Start Roll Number", "End Roll Number", "Counselor Email"])
    sheet.append(["23BQ1A5401", "23BQ1A5410", "ravindra@vvit.net"])
    sheet.append(["23BQ1A5411", "23BQ1A5420", "srinivas@vvit.net"])
    sheet.append(["23BQ1A5421", "23BQ1A5430", "kumar@vvitu.ac.in"])

    _style_sheet(sheet, 3, [22, 22, 30])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_filename(prefix: str, batch: MembershipImportBatch, ext: str) -> str:
    """Generate a filename for downloads."""
    ts = (batch.completed_at or batch.created_at).strftime("%Y%m%d_%H%M")
    return f"{prefix}_{ts}.{ext}"
