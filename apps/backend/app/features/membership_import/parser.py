"""Excel parser for the three-column membership import format.

Reads the uploaded file, locates the header row, and produces a list of
``ParsedMembershipRow`` objects.  No business logic — just parsing.
"""
from __future__ import annotations

import io
import logging
from typing import List, Tuple

from app.core.exceptions import ValidationError
from app.features.membership_import.schemas import ParsedMembershipRow
from app.features.membership_import.validation_service import validate_structure

logger = logging.getLogger("app.membership_import.parser")

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv")
MAX_UPLOAD_BYTES = 12 * 1024 * 1024


def parse_membership_excel(
    filename: str,
    content: bytes,
) -> Tuple[List[ParsedMembershipRow], List[str]]:
    """Parse an uploaded Excel into membership rows.

    Returns (rows, structural_errors).  If structural_errors is non-empty
    the file cannot be processed at all.
    """
    if not content:
        raise ValidationError("The uploaded file is empty.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValidationError(
            f"File is {len(content) / 1_048_576:.1f} MB, above the "
            f"{MAX_UPLOAD_BYTES // 1_048_576} MB limit."
        )

    grid = _load_grid(filename, content)
    if not grid:
        raise ValidationError("This file is empty.")

    headers, header_index = _find_header(grid)
    if headers is None:
        raise ValidationError(
            "Could not find a header row with 'Start Roll Number', 'End Roll Number', "
            "and 'Counselor Email' columns.  Check the file format."
        )

    column_map, structural_errors = validate_structure(headers)
    if structural_errors:
        return [], structural_errors

    rows: List[ParsedMembershipRow] = []
    for offset, raw_row in enumerate(grid[header_index + 1:], start=header_index + 2):
        cells = [_cell_to_text(c) for c in raw_row]

        # Skip blank rows
        if not any(c.strip() for c in cells):
            continue

        start_col = column_map["start roll number"]
        end_col = column_map["end roll number"]
        email_col = column_map["counselor email"]

        start_roll = cells[start_col].strip() if start_col < len(cells) else ""
        end_roll = cells[end_col].strip() if end_col < len(cells) else ""
        counselor_email = cells[email_col].strip() if email_col < len(cells) else ""

        rows.append(ParsedMembershipRow(
            row_number=offset,
            start_roll=start_roll,
            end_roll=end_roll,
            counselor_email=counselor_email,
        ))

    if not rows:
        raise ValidationError("No data rows found underneath the header row.")

    logger.info("Parsed %d rows from '%s'", len(rows), filename)
    return rows, []


# --------------------------------------------------------------------------
# File loading
# --------------------------------------------------------------------------

def _cell_to_text(value) -> str:
    """Coerce any cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _load_grid(filename: str, content: bytes) -> List[List]:
    """Load the file into a 2D grid of raw values."""
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return _load_csv(content)
    if lowered.endswith(".xls"):
        return _load_xls(content)
    if lowered.endswith((".xlsx", ".xlsm")):
        return _load_xlsx(content)
    raise ValidationError(
        f"Unsupported file type. Upload one of: {', '.join(SUPPORTED_EXTENSIONS)}."
    )


def _load_xlsx(content: bytes) -> List[List]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("Excel support unavailable (openpyxl not installed).") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(f"Could not open as Excel workbook: {exc}") from exc

    sheet = wb.active
    grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    wb.close()
    return grid


def _load_xls(content: bytes) -> List[List]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValidationError(
            "Cannot read legacy .xls format. Re-save as .xlsx."
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise ValidationError(f"Could not open as Excel workbook: {exc}") from exc

    sheet = book.sheet_by_index(0)
    grid: List[List] = []
    for row_idx in range(sheet.nrows):
        grid.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])
    return grid


def _load_csv(content: bytes) -> List[List]:
    import csv

    text = None
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:
        raise ValidationError("Could not determine CSV text encoding.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


# --------------------------------------------------------------------------
# Header detection
# --------------------------------------------------------------------------

_HEADER_SEARCH_DEPTH = 15


def _find_header(grid: List[List]) -> Tuple[List[str] | None, int]:
    """Scan the first few rows for a header containing all three expected columns."""
    from app.features.membership_import.validation_service import EXPECTED_COLUMN_ALIASES

    for idx, row in enumerate(grid[:_HEADER_SEARCH_DEPTH]):
        headers = [_cell_to_text(c).strip().lower() for c in row]
        matched = 0
        for _canonical, aliases in EXPECTED_COLUMN_ALIASES.items():
            if any(h in aliases for h in headers):
                matched += 1
        if matched >= 3:
            return [_cell_to_text(c).strip() for c in row], idx

    return None, -1
