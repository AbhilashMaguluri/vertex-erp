"""Downloadable artefacts: the credentials workbook, the import report and the
sample office template.

Everything is built in memory and streamed. Nothing is written to disk, so
there is no second copy of a credential sheet sitting in the uploads directory
waiting to be forgotten about.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, List, Optional, Sequence

from app.features.imports.models import ImportBatch
from app.features.imports.schemas import GeneratedCredential, ImportRecordResult

_HEADER_FILL = "2B3A66"   # VVIT navy
_ACCENT_FILL = "E1634B"   # VVIT coral


def _style_sheet(worksheet, column_count: int, widths: Sequence[int]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for index in range(1, column_count + 1):
        cell = worksheet.cell(row=1, column=index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.column_dimensions[get_column_letter(index)].width = (
            widths[index - 1] if index - 1 < len(widths) else 18
        )
    worksheet.row_dimensions[1].height = 22
    # Frozen header + autofilter: these sheets are read by someone scrolling
    # through several hundred rows looking for one student.
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def build_credentials_workbook(batch: ImportBatch, credentials: List[GeneratedCredential]) -> bytes:
    """Two sheets — students and counsellors — of issued logins."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()

    students = [c for c in credentials if c.record_type == "STUDENT"]
    counsellors = [c for c in credentials if c.record_type == "COUNSELLOR"]

    sheet = workbook.active
    sheet.title = "Students"
    sheet.append(["Roll Number", "Student Name", "Username", "Temporary Password", "Login Email", "Counsellor", "Status"])
    for row in students:
        sheet.append(
            [row.identifier, row.full_name, row.username, row.temporary_password, row.email, row.counsellor or "—", row.status]
        )
    _style_sheet(sheet, 7, [18, 26, 18, 20, 30, 26, 12])
    for index in range(2, len(students) + 2):
        sheet.cell(row=index, column=4).font = Font(name="Consolas", bold=True)

    staff = workbook.create_sheet("Counsellors")
    staff.append(["Counsellor (as written)", "Full Name", "Username", "Temporary Password", "Login Email", "Status"])
    for row in counsellors:
        staff.append([row.identifier, row.full_name, row.username, row.temporary_password, row.email, row.status])
    _style_sheet(staff, 6, [28, 26, 20, 20, 30, 12])
    for index in range(2, len(counsellors) + 2):
        staff.cell(row=index, column=4).font = Font(name="Consolas", bold=True)

    notes = workbook.create_sheet("Read Me")
    notes.append(["SCMS — Office Import credentials"])
    notes.append([f"Source file: {batch.original_filename}"])
    notes.append([f"Imported: {batch.completed_at.isoformat() if batch.completed_at else '—'}"])
    notes.append([])
    notes.append(["Every password on this sheet is temporary."])
    notes.append(["Each account is required to set a new password at first login."])
    notes.append(["Users sign in with their username (a student's is their roll number) or their login email."])
    notes.append(["Distribute this file over a private channel, then purge the stored credentials from"])
    notes.append(["the import's Completed screen so they are no longer held in the system."])
    notes["A1"].font = Font(bold=True, size=14, color=_ACCENT_FILL)
    notes.column_dimensions["A"].width = 96

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_report_workbook(
    batch: ImportBatch, records: List[ImportRecordResult], warnings: Sequence[str], errors: Sequence[str]
) -> bytes:
    """The import report: what the file said, what happened, and why."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()

    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    for label, value in _summary_rows(batch):
        summary.append([label, value])
    _style_sheet(summary, 2, [34, 46])

    detail = workbook.create_sheet("Records")
    detail.append(["Row", "Type", "Identifier", "Name", "Outcome", "Detail"])
    status_fills = {
        "CREATED": PatternFill("solid", fgColor="E7F6EF"),
        "REUSED": PatternFill("solid", fgColor="E8F1FB"),
        "SKIPPED": PatternFill("solid", fgColor="FDF3E2"),
        "FAILED": PatternFill("solid", fgColor="FCE9E9"),
    }
    for record in records:
        detail.append(
            [
                record.source_row_number or "—",
                record.record_type.title(),
                record.identifier,
                record.display_name or "—",
                record.status.title(),
                record.message or "",
            ]
        )
        fill = status_fills.get(record.status)
        if fill:
            detail.cell(row=detail.max_row, column=5).fill = fill
    _style_sheet(detail, 6, [8, 14, 20, 26, 14, 70])

    issues = workbook.create_sheet("Warnings & Errors")
    issues.append(["Severity", "Message"])
    for message in errors:
        issues.append(["Error", message])
    for message in warnings:
        issues.append(["Warning", message])
    if not errors and not warnings:
        issues.append(["—", "No warnings or errors were raised by this import."])
    _style_sheet(issues, 2, [14, 110])
    for index in range(2, issues.max_row + 1):
        if issues.cell(row=index, column=1).value == "Error":
            issues.cell(row=index, column=1).font = Font(bold=True, color="B42318")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_report_pdf(
    batch: ImportBatch, records: List[ImportRecordResult], warnings: Sequence[str], errors: Sequence[str]
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
        title=f"SCMS Import Report — {batch.original_filename}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ImportTitle", parent=styles["Title"], fontSize=17, textColor=colors.HexColor("#2B3A66"))
    small = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=8, leading=10)

    elements: List[Any] = [
        Paragraph("SCMS — Office Import Report", title_style),
        Paragraph(
            f"<b>{batch.original_filename}</b> &nbsp;·&nbsp; imported by "
            f"{batch.imported_by.full_name if batch.imported_by else 'Unknown'} &nbsp;·&nbsp; "
            f"{batch.completed_at.strftime('%d %b %Y, %H:%M UTC') if batch.completed_at else 'not completed'}",
            styles["BodyText"],
        ),
        Spacer(1, 8),
    ]

    summary_table = Table(
        [["Metric", "Value"]] + [[label, str(value)] for label, value in _summary_rows(batch)],
        colWidths=[70 * mm, 60 * mm],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B3A66")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6DAE5")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FB")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements += [summary_table, Spacer(1, 12), Paragraph("<b>Records</b>", styles["Heading4"])]

    # Bounded: a report PDF is read, not archived. The Excel report carries the
    # complete list when an import runs to several hundred rows.
    shown = records[:400]
    detail_rows = [["Row", "Type", "Identifier", "Name", "Outcome", "Detail"]] + [
        [
            str(r.source_row_number or "—"),
            r.record_type.title(),
            r.identifier,
            Paragraph((r.display_name or "—")[:60], small),
            r.status.title(),
            Paragraph((r.message or "")[:200], small),
        ]
        for r in shown
    ]
    detail_table = Table(
        detail_rows, colWidths=[13 * mm, 22 * mm, 34 * mm, 52 * mm, 22 * mm, 108 * mm], repeatRows=1
    )
    detail_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B3A66")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D6DAE5")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FC")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(detail_table)
    if len(records) > len(shown):
        elements += [
            Spacer(1, 6),
            Paragraph(
                f"Showing the first {len(shown)} of {len(records)} records — "
                f"download the Excel report for the complete list.",
                small,
            ),
        ]

    if errors or warnings:
        elements += [Spacer(1, 12), Paragraph("<b>Warnings &amp; Errors</b>", styles["Heading4"])]
        for message in list(errors)[:60]:
            elements.append(Paragraph(f"<font color='#B42318'><b>Error</b></font> — {message}", small))
        for message in list(warnings)[:60]:
            elements.append(Paragraph(f"<font color='#B54708'><b>Warning</b></font> — {message}", small))

    document.build(elements)
    return buffer.getvalue()


def _summary_rows(batch: ImportBatch) -> List[tuple]:
    return [
        ("Status", batch.status.title()),
        ("Source file", batch.original_filename),
        ("Imported by", batch.imported_by.full_name if batch.imported_by else "—"),
        ("Started", batch.started_at.isoformat(timespec="seconds") if batch.started_at else "—"),
        ("Completed", batch.completed_at.isoformat(timespec="seconds") if batch.completed_at else "—"),
        ("Rows in file", batch.total_rows),
        ("Students detected", batch.students_detected),
        ("Counsellors detected", batch.counsellors_detected),
        ("Students created", batch.students_created),
        ("Duplicate students skipped", batch.students_skipped),
        ("Counsellors created", batch.counsellors_created),
        ("Existing counsellors reused", batch.counsellors_reused),
        ("Counsellor assignments made", batch.assignments_created),
        ("Failed records", batch.failed_records),
        ("Warnings", batch.warning_count),
    ]


def build_sample_template() -> bytes:
    """The sample office template — deliberately the format the office already
    uses, not a technical one. Nothing on it is an identifier."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Counsellor Allotment"

    sheet.append(["S.No", "Student Roll Numbers", "Counselor Name", "Counselor Mobile"])
    sheet.append([1, "23BQ1A5401 to 23BQ1A5410", "Dr. S. Ravindra", "9440053880"])
    sheet.append([2, "23BQ1A5411 to 23BQ1A5420", "Dr. K. Satheesh", "9949397532"])
    sheet.append([3, "23BQ1A5421-5430", "Prof. Lakshmi Prasanna", "9848012345"])
    _style_sheet(sheet, 4, [8, 34, 28, 20])

    notes = workbook.create_sheet("How this works")
    lines = [
        ("SCMS — Office Import template", True),
        ("", False),
        ("Upload the allotment list your office already keeps. Nothing needs editing first.", False),
        ("", False),
        ("Required", True),
        ("  •  A column of student roll numbers. Ranges are understood:", False),
        ("        23BQ1A5401 to 23BQ1A5410      23BQ1A5401-5410", False),
        ("        23BQ1A5401 – 23BQ1A5410       23BQ1A5401 upto 5410", False),
        ("        Single roll numbers and comma-separated lists work too.", False),
        ("", False),
        ("Recognised, if present", True),
        ("  •  Counsellor Name, Counsellor Mobile, Counsellor Email", False),
        ("  •  Department / Branch, Branch Code, Section, Semester, Batch, Academic Year", False),
        ("  •  Student Name, Gender, Email, Phone, Parent Phone, Date of Birth", False),
        ("", False),
        ("Ignored", True),
        ("  •  S.No, and any other column not listed above.", False),
        ("  •  Title rows above the header — the header is found automatically.", False),
        ("", False),
        ("Anything missing is asked for once, on the Configure step, and applied to every student.", False),
    ]
    for text, bold in lines:
        notes.append([text])
        if bold:
            notes.cell(row=notes.max_row, column=1).font = Font(bold=True, color=_HEADER_FILL, size=12)
    notes.column_dimensions["A"].width = 96
    notes["A1"].font = Font(bold=True, size=15, color=_ACCENT_FILL)
    notes["A1"].alignment = Alignment(vertical="center")
    notes.row_dimensions[1].height = 24

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def download_filename(prefix: str, batch: Optional[ImportBatch], extension: str) -> str:
    stamp = (batch.completed_at or batch.created_at) if batch else datetime.utcnow()
    return f"{prefix}_{stamp.strftime('%Y%m%d_%H%M')}.{extension}"
