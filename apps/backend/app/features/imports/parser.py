"""Office file parser — reads the spreadsheet as the office actually sends it.

Two things make an office sheet awkward for a machine: the header is rarely on
the first row (there is usually a college name, a title and a blank line above
it), and the columns are named by whoever typed them ("Counselor Name",
"Name of the Counsellor", "Faculty"). This module locates the header row by
scoring candidates against a column vocabulary, maps each column onto a known
field, and ignores everything it does not recognise — including S.No.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from app.core.exceptions import ValidationError

from app.services.roll_number import MAX_ROLLS_PER_FILE, ExpansionResult, expand_roll_cell

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv")

# How far down the sheet to look for the header row. Office files put at most a
# few title/blank rows above it; scanning further only risks matching a data row.
HEADER_SEARCH_DEPTH = 25

# The fields we understand. Order matters: the most specific field wins, so
# "Counselor Mobile" never lands on the student's phone column.
#
# Each entry is (field, aliases). An alias matches when it appears as a whole
# phrase inside the normalised header; the longest matching alias decides.
COLUMN_VOCABULARY: List[Tuple[str, List[str]]] = [
    ("serial", ["s no", "sno", "sl no", "slno", "sr no", "si no", "serial no", "serial number", "serial", "sequence"]),
    (
        "roll_range",
        [
            "student roll numbers", "student roll number", "students roll numbers", "roll numbers",
            "roll number range", "roll no range", "range of roll numbers", "roll number", "roll nos",
            "roll no", "rollno", "roll", "registration numbers", "registration number", "regd nos",
            "regd no", "reg no", "hall ticket numbers", "hall ticket number", "hall ticket", "ht no",
            "htno", "student ids", "student id",
        ],
    ),
    (
        "counsellor_email",
        ["counselor email", "counsellor email", "counselor mail", "counsellor mail", "faculty email",
         "mentor email", "advisor email", "proctor email"],
    ),
    (
        "counsellor_phone",
        [
            "counselor mobile", "counsellor mobile", "counselor phone", "counsellor phone",
            "counselor contact", "counsellor contact", "counselor number", "counsellor number",
            "faculty mobile", "faculty phone", "faculty contact", "mentor mobile", "mentor phone",
            "advisor mobile", "advisor phone", "proctor mobile", "proctor phone",
        ],
    ),
    (
        "counsellor_name",
        [
            "counselor name", "counsellor name", "name of the counselor", "name of the counsellor",
            "name of counselor", "name of counsellor", "faculty name", "name of the faculty",
            "mentor name", "advisor name", "proctor name", "class teacher", "counselor", "counsellor",
            "mentor", "advisor", "proctor", "faculty",
        ],
    ),
    ("parent_phone", ["parent phone", "parent mobile", "parent contact", "parent number", "father mobile",
                      "father phone", "mother mobile", "mother phone", "guardian phone", "guardian mobile"]),
    ("student_email", ["student email", "student mail", "email id", "mail id", "e mail", "email", "mail"]),
    ("student_phone", ["student mobile", "student phone", "student contact", "mobile number",
                       "phone number", "contact number"]),
    ("student_name", ["student name", "name of the student", "name of student", "students name"]),
    ("branch_code", ["branch code", "department code", "dept code"]),
    ("department", ["department name", "department", "dept", "branch name", "branch", "programme", "program"]),
    ("academic_year", ["academic year", "acad year", "academic session", "a y"]),
    ("semester", ["semester", "sem"]),
    ("section", ["section", "sec"]),
    ("batch", ["batch year", "batch", "year of admission", "admission year", "admitted year"]),
    ("gender", ["gender", "sex"]),
    ("date_of_birth", ["date of birth", "dob", "birth date", "d o b"]),
    # Deliberately last: a bare "mobile"/"phone"/"name" is only assigned once
    # every more specific reading has been ruled out. Post-processing decides
    # whether an ambiguous phone belongs to the counsellor or the student.
    ("ambiguous_phone", ["mobile", "phone", "contact", "cell"]),
    ("ambiguous_name", ["name"]),
]

_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def normalise_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = _NON_ALNUM.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


def _match_field(header: str) -> Tuple[Optional[str], int]:
    """Best (field, alias-length) for one header cell. Longer alias == more
    specific reading, which is what breaks ties between overlapping vocabularies."""
    normalised = normalise_header(header)
    if not normalised:
        return None, 0

    best_field: Optional[str] = None
    best_len = 0
    for field_name, aliases in COLUMN_VOCABULARY:
        for alias in aliases:
            if normalised == alias:
                # An exact header match is authoritative; nothing beats it.
                return field_name, len(alias) + 100
            if re.search(rf"(?:^|\s){re.escape(alias)}(?:\s|$)", normalised):
                if len(alias) > best_len:
                    best_field, best_len = field_name, len(alias)
    return best_field, best_len


@dataclass
class ParsedRow:
    """One row of the office sheet, already expanded."""

    row_number: int
    roll_numbers: List[str] = field(default_factory=list)
    range_segments: List[str] = field(default_factory=list)
    raw_roll_text: str = ""
    counsellor_name: Optional[str] = None
    counsellor_phone: Optional[str] = None
    counsellor_email: Optional[str] = None
    student_name: Optional[str] = None
    department: Optional[str] = None
    branch_code: Optional[str] = None
    academic_year: Optional[str] = None
    semester: Optional[str] = None
    section: Optional[str] = None
    batch: Optional[str] = None
    gender: Optional[str] = None
    student_email: Optional[str] = None
    student_phone: Optional[str] = None
    parent_phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return bool(self.roll_numbers) and not self.errors


@dataclass
class ParsedFile:
    sheet_name: str
    header_row_number: int
    detected_columns: Dict[str, str] = field(default_factory=dict)  # field -> source header text
    ignored_columns: List[str] = field(default_factory=list)
    rows: List[ParsedRow] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# --------------------------------------------------------------------------
# Sheet loading
# --------------------------------------------------------------------------

def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        # Excel stores every number as a float; "2023.0" must read back as the
        # batch year 2023, not as a decimal.
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _load_xlsx(content: bytes) -> Tuple[str, List[List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise ValidationError("Excel support is unavailable on the server (openpyxl is not installed).") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — any openpyxl failure means "not a readable workbook"
        raise ValidationError(f"This file could not be opened as an Excel workbook: {exc}") from exc

    sheet = workbook.active
    grid = [[_cell_to_text(c) for c in row] for row in sheet.iter_rows(values_only=True)]
    name = sheet.title
    workbook.close()
    return name, grid


def _load_xls(content: bytes) -> Tuple[str, List[List[str]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValidationError(
            "This server cannot read the legacy .xls format. Re-save the file as .xlsx or .csv "
            "and upload it again."
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:  # noqa: BLE001
        raise ValidationError(f"This file could not be opened as an Excel workbook: {exc}") from exc

    sheet = book.sheet_by_index(0)
    grid: List[List[str]] = []
    for row_index in range(sheet.nrows):
        row: List[str] = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                row.append(_cell_to_text(datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))))
            else:
                row.append(_cell_to_text(cell.value))
        grid.append(row)
    return sheet.name, grid


def _load_csv(content: bytes) -> Tuple[str, List[List[str]]]:
    text: Optional[str] = None
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:  # pragma: no cover — latin-1 decodes any byte string
        raise ValidationError("The text encoding of this CSV file could not be determined.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    return "CSV", [[_cell_to_text(cell) for cell in row] for row in reader]


def load_grid(filename: str, content: bytes) -> Tuple[str, List[List[str]]]:
    """Read any supported office file into a plain rectangular grid of strings."""
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return _load_csv(content)
    if lowered.endswith(".xls"):
        return _load_xls(content)
    if lowered.endswith((".xlsx", ".xlsm")):
        return _load_xlsx(content)
    raise ValidationError(
        f"Unsupported file type. Upload one of: {', '.join(SUPPORTED_EXTENSIONS)}."
    )


# --------------------------------------------------------------------------
# Header & column detection
# --------------------------------------------------------------------------

def _score_header_row(row: Sequence[str]) -> Tuple[int, Dict[int, str]]:
    """How strongly this row reads as a header, plus the column map it implies."""
    mapping: Dict[int, str] = {}
    score = 0
    for index, cell in enumerate(row):
        field_name, weight = _match_field(cell)
        if field_name:
            mapping[index] = field_name
            score += weight
    # A header row must at least tell us which column holds the students.
    if "roll_range" not in mapping.values():
        return 0, {}
    return score, mapping


def detect_header(grid: List[List[str]]) -> Tuple[int, Dict[int, str]]:
    best_index, best_score, best_map = -1, 0, {}
    for index, row in enumerate(grid[:HEADER_SEARCH_DEPTH]):
        score, mapping = _score_header_row(row)
        if score > best_score:
            best_index, best_score, best_map = index, score, mapping

    if best_index < 0:
        raise ValidationError(
            "No student roll-number column could be found in this file. The sheet needs a column "
            "headed something like 'Student Roll Numbers' — everything else is optional."
        )
    return best_index, best_map


def _resolve_ambiguous(mapping: Dict[int, str]) -> Dict[int, str]:
    """Decide what a bare 'Mobile' or 'Name' column actually refers to.

    On a range sheet every row describes a group of students under one member
    of staff, so an unqualified contact column belongs to that member of staff.
    A bare 'Name' is only read as the counsellor when no counsellor column was
    found by name; otherwise it is the student's.
    """
    assigned = set(mapping.values())
    resolved = dict(mapping)
    for index, field_name in mapping.items():
        if field_name == "ambiguous_phone":
            if "counsellor_phone" not in assigned:
                resolved[index] = "counsellor_phone"
                assigned.add("counsellor_phone")
            elif "student_phone" not in assigned:
                resolved[index] = "student_phone"
                assigned.add("student_phone")
            else:
                resolved.pop(index)
        elif field_name == "ambiguous_name":
            if "counsellor_name" not in assigned:
                resolved[index] = "counsellor_name"
                assigned.add("counsellor_name")
            elif "student_name" not in assigned:
                resolved[index] = "student_name"
                assigned.add("student_name")
            else:
                resolved.pop(index)
    return resolved


# --------------------------------------------------------------------------
# Row extraction
# --------------------------------------------------------------------------

_ROW_TEXT_FIELDS = (
    "counsellor_name", "counsellor_phone", "counsellor_email", "student_name", "department",
    "branch_code", "academic_year", "semester", "section", "batch", "gender", "student_email",
    "student_phone", "parent_phone", "date_of_birth",
)


def _clean_phone(value: str) -> Optional[str]:
    """Keep the digits (and a leading +) an office phone column contains.
    Anything under 10 digits is not a dialable Indian number and is dropped."""
    if not value:
        return None
    digits = re.sub(r"\D", "", value)
    if len(digits) < 10:
        return None
    prefix = "+" if value.lstrip().startswith("+") else ""
    return f"{prefix}{digits}"[:20]


def parse_office_file(filename: str, content: bytes) -> ParsedFile:
    """Read an office sheet into expanded, per-row import instructions."""
    sheet_name, grid = load_grid(filename, content)
    if not grid:
        raise ValidationError("This file is empty.")

    header_index, raw_mapping = detect_header(grid)
    mapping = _resolve_ambiguous(raw_mapping)

    header_row = grid[header_index]
    detected_columns: Dict[str, str] = {}
    for index, field_name in mapping.items():
        if field_name == "serial":
            continue  # S.No carries no information — ignored by design.
        source = header_row[index] if index < len(header_row) else f"Column {index + 1}"
        detected_columns.setdefault(field_name, str(source).strip())

    ignored_columns = [
        str(cell).strip()
        for index, cell in enumerate(header_row)
        if str(cell).strip() and (index not in mapping or mapping[index] == "serial")
    ]

    parsed = ParsedFile(
        sheet_name=sheet_name,
        header_row_number=header_index + 1,
        detected_columns=detected_columns,
        ignored_columns=ignored_columns,
    )

    roll_column = next((i for i, f in mapping.items() if f == "roll_range"), None)
    if roll_column is None:  # pragma: no cover — detect_header already guarantees one
        raise ValidationError("No student roll-number column could be found in this file.")

    total_rolls = 0
    for offset, raw_row in enumerate(grid[header_index + 1 :], start=header_index + 2):
        cells = [str(c).strip() for c in raw_row]
        if not any(cells):
            continue  # blank spacer row

        row = ParsedRow(row_number=offset)
        for index, field_name in mapping.items():
            if field_name in ("serial", "roll_range"):
                continue
            value = cells[index].strip() if index < len(cells) else ""
            if not value:
                continue
            if field_name in ("counsellor_phone", "student_phone", "parent_phone"):
                setattr(row, field_name, _clean_phone(value))
            elif field_name in _ROW_TEXT_FIELDS:
                setattr(row, field_name, value)

        raw_roll = cells[roll_column] if roll_column < len(cells) else ""
        row.raw_roll_text = raw_roll
        if not raw_roll:
            # A row with staff details but no students is a footer/notes line.
            if not any(getattr(row, f) for f in _ROW_TEXT_FIELDS):
                continue
            row.errors.append("no roll number in this row")
            parsed.rows.append(row)
            continue

        expansion: ExpansionResult = expand_roll_cell(raw_roll)
        row.roll_numbers = expansion.roll_numbers
        row.range_segments = expansion.segments
        row.warnings.extend(expansion.warnings)
        row.errors.extend(expansion.errors)
        total_rolls += len(expansion.roll_numbers)
        parsed.rows.append(row)

    if not parsed.rows:
        raise ValidationError("No data rows were found underneath the header of this file.")

    if total_rolls > MAX_ROLLS_PER_FILE:
        raise ValidationError(
            f"This file expands to {total_rolls:,} students, above the {MAX_ROLLS_PER_FILE:,} per-import "
            f"limit. Split it into smaller files (one department or batch each) and import them in turn."
        )

    if "counsellor_name" not in detected_columns:
        parsed.warnings.append(
            "No counsellor column was recognised — students will be created without a counsellor assignment."
        )
    return parsed
