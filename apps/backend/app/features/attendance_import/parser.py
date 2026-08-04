"""Excel parser for the two-column attendance import format."""
from __future__ import annotations

import io
import logging
from typing import List, Tuple

from app.core.exceptions import ValidationError
from app.features.attendance_import.schemas import ParsedAttendanceRow
from app.features.attendance_import.validation_service import validate_structure

logger = logging.getLogger("app.attendance_import.parser")

SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv")
MAX_UPLOAD_BYTES = 12 * 1024 * 1024


def parse_attendance_excel(
    filename: str,
    content: bytes,
) -> Tuple[List[ParsedAttendanceRow], List[str]]:
    """Parse an uploaded attendance Excel file into row objects."""
    if not content:
        raise ValidationError("The uploaded file is empty.")
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValidationError(
            f"File is {len(content) / 1_048_576:.1f} MB, above the "
            f"{MAX_UPLOAD_BYTES // 1_048_576} MB limit."
        )

    grid = _load_grid(filename, content)
    if not grid:
        raise ValidationError("This file is empty.")

    headers, header_index = _find_header(grid)
    if headers is None:
        raise ValidationError(
            "Could not find a header row with 'Student Roll Number' and 'Attendance Status' "
            "columns. Please check your Excel headers."
        )

    column_map, structural_errors = validate_structure(headers)
    if structural_errors:
        return [], structural_errors

    rows: List[ParsedAttendanceRow] = []
    roll_col = column_map["student_roll"]
    status_col = column_map["attendance_status"]

    for offset, raw_row in enumerate(grid[header_index + 1:], start=header_index + 2):
        cells = [_cell_to_text(c) for c in raw_row]

        # Skip blank spacer rows
        if not any(c.strip() for c in cells):
            continue

        student_roll = cells[roll_col].strip() if roll_col < len(cells) else ""
        raw_status = cells[status_col].strip() if status_col < len(cells) else ""

        rows.append(ParsedAttendanceRow(
            row_number=offset,
            student_roll=student_roll,
            raw_status=raw_status,
        ))

    if not rows:
        raise ValidationError("No data rows found underneath the header row.")

    logger.info("Parsed %d attendance rows from '%s'", len(rows), filename)
    return rows, []


# --------------------------------------------------------------------------
# File loading helpers
# --------------------------------------------------------------------------

def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _load_grid(filename: str, content: bytes) -> List[List]:
    lowered = (filename or "").lower()
    if lowered.endswith(".csv"):
        return _load_csv(content)
    if lowered.endswith(".xls"):
        return _load_xls(content)
    if lowered.endswith((".xlsx", ".xlsm")):
        return _load_xlsx(content)
    raise ValidationError(
        f"Unsupported file type. Upload one of: {', '.join(SUPPORTED_EXTENSIONS)}."
    )


def _load_xlsx(content: bytes) -> List[List]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("Excel support unavailable (openpyxl not installed).") from exc

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(f"Could not open as Excel workbook: {exc}") from exc

    sheet = wb.active
    grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    wb.close()
    return grid


def _load_xls(content: bytes) -> List[List]:
    try:
        import xlrd
    except ImportError as exc:
        raise ValidationError("Cannot read legacy .xls format. Re-save as .xlsx.") from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise ValidationError(f"Could not open as Excel workbook: {exc}") from exc

    sheet = book.sheet_by_index(0)
    grid: List[List] = []
    for row_idx in range(sheet.nrows):
        grid.append([sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)])
    return grid


def _load_csv(content: bytes) -> List[List]:
    import csv

    text = None
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:
        raise ValidationError("Could not determine CSV text encoding.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


# --------------------------------------------------------------------------
# Header scanning
# --------------------------------------------------------------------------

_HEADER_SEARCH_DEPTH = 15


def _find_header(grid: List[List]) -> Tuple[List[str] | None, int]:
    from app.features.attendance_import.validation_service import EXPECTED_COLUMN_ALIASES

    for idx, row in enumerate(grid[:_HEADER_SEARCH_DEPTH]):
        headers = [_cell_to_text(c).strip().lower() for c in row]
        matched = 0
        for _canonical, aliases in EXPECTED_COLUMN_ALIASES.items():
            if any(h in aliases for h in headers):
                matched += 1
        if matched >= 2:
            return [_cell_to_text(c).strip() for c in row], idx

    return None, -1
