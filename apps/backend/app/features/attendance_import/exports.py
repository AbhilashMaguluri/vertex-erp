"""Downloadable artefacts for the Attendance Import feature.

Generates:
- Attendance error report (Attendance_Import_Errors.xlsx)
- Attendance import report (Attendance_Import_Report.xlsx)
- Sample template (Attendance_Import_Template.xlsx)
"""
from __future__ import annotations

import io
from typing import List, Sequence

from app.features.attendance_import.models import AttendanceImportBatch
from app.features.attendance_import.schemas import (
    AttendanceImportResultRecord,
    ValidationErrorRow,
)

_HEADER_FILL = "2B3A66"


def _style_sheet(worksheet, column_count: int, widths: Sequence[int]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for idx in range(1, column_count + 1):
        cell = worksheet.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.column_dimensions[get_column_letter(idx)].width = (
            widths[idx - 1] if idx - 1 < len(widths) else 18
        )
    worksheet.row_dimensions[1].height = 22
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def build_error_report(validation_errors: List[ValidationErrorRow]) -> bytes:
    """Build Attendance_Import_Errors.xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Import Errors"

    sheet.append(["Row Number", "Student Roll Number", "Error", "Suggested Fix"])

    for err in validation_errors:
        sheet.append([err.row, err.roll_number, err.error, err.suggested_fix])

    _style_sheet(sheet, 4, [12, 22, 50, 40])

    error_font = Font(color="CC0000")
    for row_idx in range(2, len(validation_errors) + 2):
        for col_idx in range(1, 5):
            sheet.cell(row=row_idx, column=col_idx).font = error_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_report_workbook(
    batch: AttendanceImportBatch,
    records: List[AttendanceImportResultRecord],
    warnings: List[str],
    errors: List[str],
) -> bytes:
    """Build detailed import report workbook."""
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Import Report"

    sheet.append(["Type", "Roll Number", "Student Name", "Status", "Message", "Row"])
    for r in records:
        sheet.append([
            r.record_type, r.identifier,
            r.display_name or "", r.status,
            r.message or "", r.source_row_number or "",
        ])
    _style_sheet(sheet, 6, [14, 22, 26, 14, 50, 8])

    if warnings or errors:
        notes = wb.create_sheet("Warnings & Errors")
        notes.append(["Type", "Message"])
        for w in warnings:
            notes.append(["WARNING", w])
        for e in errors:
            notes.append(["ERROR", e])
        _style_sheet(notes, 2, [12, 80])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_sample_template() -> bytes:
    """Build the sample two-column attendance template."""
    from openpyxl import Workbook

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Attendance Import"

    sheet.append(["Student Roll Number", "Attendance Status"])
    sheet.append(["23BQ1A5401", "Present"])
    sheet.append(["23BQ1A5402", "Absent"])
    sheet.append(["23BQ1A5403", "P"])
    sheet.append(["23BQ1A5404", "A"])
    sheet.append(["23BQ1A5405", "OD"])
    sheet.append(["23BQ1A5406", "ML"])

    _style_sheet(sheet, 2, [24, 24])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def download_filename(prefix: str, batch: AttendanceImportBatch, ext: str) -> str:
    ts = (batch.completed_at or batch.created_at).strftime("%Y%m%d_%H%M")
    return f"{prefix}_{ts}.{ext}"
